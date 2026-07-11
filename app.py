import io
import base64
from math import ceil

import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from PIL import Image
import pandas as pd
from datetime import date

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

EMPLOYEES_SHEET = "Employees"
ATTENDANCE_SHEET = "Attendance"
EMPLOYEE_HEADERS = ["id", "name", "photo_url", "division", "order"]
# Posisi kolom tetap (1-indexed) - dihardcode biar gak perlu baca header berulang-ulang ke API
COL = {"id": 1, "name": 2, "photo_url": 3, "division": 4, "order": 5}

STATUS_CYCLE = ["belum", "hijau", "merah", "kuning"]
STATUS_ICON = {"belum": "\u26AA", "hijau": "\U0001F7E2", "merah": "\U0001F534", "kuning": "\U0001F7E1"}
STATUS_LABEL = {
    "belum": "Belum diisi",
    "hijau": "Tepat waktu (\u2264 08.00)",
    "merah": "Telat (> 08.00)",
    "kuning": "Tidak hadir (sakit/cuti/SPD/izin)",
}
STATUS_COLORS = {"hijau": "#2F9E56", "merah": "#E14B3F", "kuning": "#DE9A1F", "belum": "#C9C7BF"}

KACAB_NAME = "Kacab"
DEFAULT_DIVISION_CONFIG = "Service:4, Marketing:3, Admin:2, Part:2"

DIVISION_RENAME_MAP = {
    "Sales": "Marketing",
}


def next_status(current):
    idx = STATUS_CYCLE.index(current)
    return STATUS_CYCLE[(idx + 1) % len(STATUS_CYCLE)]


def parse_division_config(raw):
    order_list = []
    cols_map = {}
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            name, cols = part.split(":", 1)
            name = name.strip()
            try:
                cols = max(1, int(cols.strip()))
            except ValueError:
                cols = 1
        else:
            name = part
            cols = 1
        if name:
            order_list.append(name)
            cols_map[name] = cols
    return order_list, cols_map


def chunk_column_major(items, n_chunks):
    n_chunks = max(1, n_chunks)
    if not items:
        return [[] for _ in range(n_chunks)]
    size = ceil(len(items) / n_chunks)
    chunks = [items[i : i + size] for i in range(0, len(items), size)]
    while len(chunks) < n_chunks:
        chunks.append([])
    return chunks[:n_chunks]


# ---------- Google Sheets connection ----------

def get_credentials():
    if "gcp_service_account" in st.secrets:
        creds_dict = dict(st.secrets["gcp_service_account"])
        return Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return Credentials.from_service_account_file("credentials.json", scopes=SCOPES)


@st.cache_resource
def get_client():
    return gspread.authorize(get_credentials())


@st.cache_resource
def get_spreadsheet():
    spreadsheet_id = st.secrets.get("SPREADSHEET_ID")
    if not spreadsheet_id:
        st.error(
            "SPREADSHEET_ID belum diisi. Tambahkan di .streamlit/secrets.toml "
            "(lihat README.md untuk caranya)."
        )
        st.stop()
    return get_client().open_by_key(spreadsheet_id)


def _ensure_headers(ws, required_headers):
    headers = ws.row_values(1)
    changed = False
    for h in required_headers:
        if h not in headers:
            headers.append(h)
            changed = True
    if changed:
        ws.update("A1", [headers])


@st.cache_resource
def get_worksheets():
    """Cache_resource: worksheet handle & migrasi header cuma dijalankan SEKALI
    per sesi server, bukan tiap kali fungsi lain dipanggil. Ini yang paling
    banyak menghemat kuota API dibanding sebelumnya."""
    ss = get_spreadsheet()
    try:
        ws_emp = ss.worksheet(EMPLOYEES_SHEET)
    except gspread.WorksheetNotFound:
        ws_emp = ss.add_worksheet(title=EMPLOYEES_SHEET, rows=300, cols=len(EMPLOYEE_HEADERS))
        ws_emp.append_row(EMPLOYEE_HEADERS)
    _ensure_headers(ws_emp, EMPLOYEE_HEADERS)

    try:
        ws_att = ss.worksheet(ATTENDANCE_SHEET)
    except gspread.WorksheetNotFound:
        ws_att = ss.add_worksheet(title=ATTENDANCE_SHEET, rows=3000, cols=3)
        ws_att.append_row(["employee_id", "date", "status"])
    return ws_emp, ws_att


MAX_DATA_URI_CHARS = 45000


def process_uploaded_photo(uploaded_file):
    img = Image.open(uploaded_file)
    img = img.convert("RGB")

    size = 220
    quality = 75
    for _ in range(6):
        resized = img.copy()
        resized.thumbnail((size, size))
        buf = io.BytesIO()
        resized.save(buf, format="JPEG", quality=quality, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode()
        data_uri = f"data:image/jpeg;base64,{b64}"
        if len(data_uri) <= MAX_DATA_URI_CHARS:
            return data_uri
        size = int(size * 0.8)
        quality = max(35, quality - 10)

    raise ValueError(
        "Foto masih terlalu besar walau sudah dikompres habis-habisan. "
        "Coba pakai foto lain (ukuran file asli lebih kecil)."
    )


# ---------- Data access (dengan cache biar hemat kuota) ----------

@st.cache_data(ttl=120, show_spinner=False)
def _fetch_employee_records():
    ws_emp, _ = get_worksheets()
    return ws_emp.get_all_records()


@st.cache_data(ttl=8, show_spinner=False)
def _fetch_attendance_records():
    _, ws_att = get_worksheets()
    return ws_att.get_all_records()


def _invalidate_employee_cache():
    _fetch_employee_records.clear()


def _invalidate_attendance_cache():
    _fetch_attendance_records.clear()


def load_employees_df():
    records = _fetch_employee_records()
    df = pd.DataFrame(records)
    if df.empty:
        df = pd.DataFrame(columns=EMPLOYEE_HEADERS)
    for col in ["photo_url", "division", "order"]:
        if col not in df.columns:
            df[col] = ""
    df["division"] = df["division"].replace("", "Lainnya").fillna("Lainnya")
    df["division"] = df["division"].replace(DIVISION_RENAME_MAP)
    df["order"] = pd.to_numeric(df["order"], errors="coerce")
    df["order"] = df["order"].fillna(df["id"])
    return df


def _find_employee_row(emp_id):
    """Cari nomor baris di sheet dari data yang sudah di-cache (tanpa API call
    tambahan), asal urutan baris di sheet tidak diubah manual dari luar."""
    records = _fetch_employee_records()
    for i, r in enumerate(records):
        if str(r["id"]) == str(emp_id):
            return i + 2
    return None


def add_employee(name, photo_url="", division="Lainnya", order=None):
    ws_emp, _ = get_worksheets()
    df = load_employees_df()
    new_id = int(df["id"].max()) + 1 if not df.empty else 1
    if order is None:
        order = new_id
    ws_emp.append_row([new_id, name, photo_url, division or "Lainnya", order])
    _invalidate_employee_cache()
    return new_id


def update_employee_name(emp_id, new_name):
    ws_emp, _ = get_worksheets()
    row = _find_employee_row(emp_id)
    if row:
        ws_emp.update_cell(row, COL["name"], new_name)
        _invalidate_employee_cache()


def update_employee_photo(emp_id, photo_url):
    ws_emp, _ = get_worksheets()
    row = _find_employee_row(emp_id)
    if row:
        ws_emp.update_cell(row, COL["photo_url"], photo_url)
        _invalidate_employee_cache()


def update_employee_division(emp_id, division):
    ws_emp, _ = get_worksheets()
    row = _find_employee_row(emp_id)
    if row:
        ws_emp.update_cell(row, COL["division"], division)
        _invalidate_employee_cache()


def update_employee_order(emp_id, order_value):
    ws_emp, _ = get_worksheets()
    row = _find_employee_row(emp_id)
    if row:
        ws_emp.update_cell(row, COL["order"], order_value)
        _invalidate_employee_cache()


def delete_employee(emp_id):
    ws_emp, ws_att = get_worksheets()
    row = _find_employee_row(emp_id)
    if row:
        ws_emp.delete_rows(row)
    records = _fetch_attendance_records()
    rows_to_delete = [i + 2 for i, r in enumerate(records) if str(r["employee_id"]) == str(emp_id)]
    for r in sorted(rows_to_delete, reverse=True):
        ws_att.delete_rows(r)
    _invalidate_employee_cache()
    _invalidate_attendance_cache()


def load_attendance_for_date(date_str):
    records = _fetch_attendance_records()
    return {str(r["employee_id"]): r["status"] for r in records if str(r["date"]) == date_str}


MONTH_NAMES_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]

# Label singkat khusus buat rekap yang di-download, biar orang lain yang buka
# langsung paham tanpa perlu tahu arti warna hijau/merah/kuning
EXPORT_STATUS_LABEL = {
    "belum": "Belum diisi",
    "hijau": "Tepat waktu",
    "merah": "Telat",
    "kuning": "Absen",
}

EXPORT_FILL_HEX = {
    "Tepat waktu": "C6EFCE",
    "Telat": "FFC7CE",
    "Absen": "FFEB9C",
    "Belum diisi": "F2F2F2",
}


def build_monthly_report(year, month):
    """Rekap kehadiran 1 bulan: baris = karyawan, kolom = tiap tanggal di bulan itu."""
    records = _fetch_attendance_records()
    if not records:
        return pd.DataFrame()
    df_att = pd.DataFrame(records)
    month_str = f"{year:04d}-{month:02d}"
    df_month = df_att[df_att["date"].astype(str).str.startswith(month_str)].copy()
    if df_month.empty:
        return pd.DataFrame()

    emp_df = load_employees_df()[["id", "name", "division"]].copy()
    emp_df["id"] = emp_df["id"].astype(str)
    df_month["employee_id"] = df_month["employee_id"].astype(str)

    merged = df_month.merge(emp_df, left_on="employee_id", right_on="id", how="left")
    merged["name"] = merged["name"].fillna("(karyawan sudah dihapus - id " + merged["employee_id"] + ")")
    merged["division"] = merged["division"].fillna("Lainnya")
    merged["status_label"] = merged["status"].map(EXPORT_STATUS_LABEL).fillna(merged["status"])

    pivot = merged.pivot_table(
        index=["division", "name"], columns="date", values="status_label", aggfunc="first"
    )
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot = pivot.fillna("Belum diisi")
    pivot = pivot.reset_index().rename(columns={"division": "Divisi", "name": "Nama"})
    return pivot


def build_excel_bytes(df, sheet_name="Rekap Kehadiran"):
    """Bikin file .xlsx asli (bukan CSV) dengan header rapi + warna sel sesuai
    status, biar bisa langsung dibuka di aplikasi apa pun yang bisa baca Excel
    (WPS, LibreOffice, Google Sheets, dst) tanpa perlu convert dulu."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name]])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 28)
            if col_name not in ("Divisi", "Nama"):
                for row_idx, value in enumerate(df[col_name], start=2):
                    hex_color = EXPORT_FILL_HEX.get(str(value))
                    if hex_color:
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color=hex_color, end_color=hex_color, fill_type="solid"
                        )
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

        ws.freeze_panes = "C2"

    buf.seek(0)
    return buf.getvalue()


def save_pending_attendance(date_str, pending_changes):
    """Simpan banyak perubahan kehadiran sekaligus dalam sedikit mungkin request:
    - update sel yang sudah ada -> 1x batch_update (untuk semuanya sekaligus)
    - baris baru -> 1x append_rows (untuk semuanya sekaligus)
    - balik ke 'belum diisi' -> delete_rows (baris demi baris, biasanya jarang terjadi)
    Jadi 75 perubahan pun tetap hanya makan beberapa request, bukan 75 request.
    """
    if not pending_changes:
        return
    _, ws_att = get_worksheets()
    records = _fetch_attendance_records()

    existing_row_for_emp = {}
    for i, r in enumerate(records):
        if str(r["date"]) == date_str:
            existing_row_for_emp[str(r["employee_id"])] = i + 2

    cell_updates = []
    new_rows = []
    rows_to_delete = []

    for emp_id, status in pending_changes.items():
        row = existing_row_for_emp.get(str(emp_id))
        if status == "belum":
            if row:
                rows_to_delete.append(row)
        else:
            if row:
                cell_updates.append({"range": f"C{row}", "values": [[status]]})
            else:
                new_rows.append([emp_id, date_str, status])

    if cell_updates:
        ws_att.batch_update(cell_updates)
    if new_rows:
        ws_att.append_rows(new_rows)
    for row in sorted(rows_to_delete, reverse=True):
        ws_att.delete_rows(row)

    _invalidate_attendance_cache()


def get_working_attendance(date_str):
    """Salinan kerja kehadiran untuk tanggal ini, disimpan di session (di memori
    browser/server sesi ini) - klik ganti status TIDAK langsung ke Google Sheets,
    baru dikirim pas tombol 'Simpan' ditekan (atau otomatis saat pindah tanggal)."""
    data_key = f"att_data_{date_str}"
    dirty_key = f"att_dirty_{date_str}"
    if data_key not in st.session_state:
        st.session_state[data_key] = dict(load_attendance_for_date(date_str))
        st.session_state[dirty_key] = set()
    return st.session_state[data_key], st.session_state[dirty_key]


def autosave_if_switching_date(new_date_str):
    prev_key = "current_working_date"
    prev_date = st.session_state.get(prev_key)
    if prev_date and prev_date != new_date_str:
        old_dirty = st.session_state.get(f"att_dirty_{prev_date}", set())
        if old_dirty:
            old_data = st.session_state.get(f"att_data_{prev_date}", {})
            save_pending_attendance(prev_date, {eid: old_data[eid] for eid in old_dirty})
            st.session_state[f"att_dirty_{prev_date}"] = set()
    st.session_state[prev_key] = new_date_str


def placeholder_photo_url(emp_id):
    idx = (int(emp_id) % 70) + 1
    return f"https://i.pravatar.cc/300?img={idx}"


# ---------- UI ----------

st.set_page_config(page_title="Board Kehadiran Karyawan", page_icon="\U0001F5C2\uFE0F", layout="wide")

st.markdown(
    """
    <style>
    .stApp { background-color: #FAFAF8; }
    .board-title { text-align: center; }
    .board-title h1 {
        font-family: 'Trebuchet MS', sans-serif; font-weight: 800;
        margin-bottom: 0; letter-spacing: 0.02em;
    }
    .board-title h3 {
        font-family: 'Trebuchet MS', sans-serif; font-weight: 600;
        margin-top: 0; color: #444; letter-spacing: 0.03em;
    }
    .legend-box {
        display: flex; gap: 22px; justify-content: center; flex-wrap: wrap;
        margin: 6px 0 18px 0; font-size: 13px;
    }
    .legend-item { display: flex; align-items: center; gap: 6px; }
    .legend-dot { width: 14px; height: 14px; border-radius: 50%; display: inline-block; }
    .division-header {
        text-align: center; font-weight: 700; font-size: 15px;
        border-bottom: 2px solid #333; padding-bottom: 6px; margin-bottom: 10px;
        text-transform: uppercase; letter-spacing: 0.05em;
    }
    .kacab-header {
        text-align: center; font-weight: 700; font-size: 15px;
        margin-bottom: 10px; text-transform: uppercase; letter-spacing: 0.05em;
    }
    .emp-row-name { font-size: 11.5px; color: #333; line-height: 1.2; text-align: center; }
    .emp-photo-wrap { text-align: center; margin-bottom: 4px; }
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("Pengaturan Board")
    board_title = st.text_input("Judul board", value="KEHADIRAN KARYAWAN")
    board_subtitle = st.text_input("Sub-judul (nama cabang)", value="CABANG WARU")
    division_config_raw = st.text_input(
        "Urutan & jumlah kolom divisi",
        value=DEFAULT_DIVISION_CONFIG,
        help="Format: NamaDivisi:JumlahKolom, pisahkan dengan koma. "
        "'Kacab' otomatis ditaruh sendiri di paling atas, tidak perlu ditulis di sini.",
    )
    show_names = st.checkbox("Tampilkan nama di kartu", value=True)

    known_divisions, division_cols_config = parse_division_config(division_config_raw)

    st.divider()
    st.header("Kelola Karyawan")
    division_options = [KACAB_NAME] + known_divisions + ["Lainnya"]

    # Data karyawan dimuat SEKALI di sini, dipakai ulang di sidebar & board utama
    employees_df = load_employees_df()

    with st.form("add_employee_form", clear_on_submit=True):
        new_name = st.text_input("Nama / label (mis. A, B, C jika belum tahu nama aslinya)")
        new_division = st.selectbox("Divisi", options=division_options)
        new_photo_file = st.file_uploader(
            "Upload foto (opsional)", type=["jpg", "jpeg", "png", "webp"]
        )
        new_photo_url = st.text_input(
            "Atau isi URL foto (kosongkan kalau upload file di atas / belum ada)",
            help="Kalau dua-duanya kosong, otomatis dikasih foto sementara (placeholder).",
        )
        submitted = st.form_submit_button("+ Tambah")
        if submitted and new_name.strip():
            photo_url = new_photo_url.strip()
            if new_photo_file is not None:
                try:
                    with st.spinner("Memproses foto..."):
                        photo_url = process_uploaded_photo(new_photo_file)
                except Exception as e:
                    st.error(f"Gagal memproses foto: {e}")
                    st.stop()
            add_employee(new_name.strip(), photo_url, new_division)
            st.rerun()

    st.divider()

    @st.fragment
    def render_employee_management_panel(emp_df, division_options):
        if emp_df.empty:
            return
        st.caption("Ubah nama / divisi / urutan / foto / hapus karyawan")
        for _, row in emp_df.sort_values(["division", "order"]).iterrows():
            with st.expander(f"{row['name']} ({row['division']})"):
                edited_name = st.text_input("Nama", value=row["name"], key=f"name_{row['id']}")
                opts = division_options if row["division"] in division_options else division_options + [row["division"]]
                edited_division = st.selectbox(
                    "Divisi", options=opts, index=opts.index(row["division"]), key=f"div_{row['id']}"
                )
                edited_order = st.number_input(
                    "Urutan (angka kecil = tampil duluan)",
                    value=float(row["order"]), step=1.0, key=f"order_{row['id']}",
                )
                csave, cdel = st.columns(2)
                if csave.button("\U0001F4BE Simpan", key=f"save_{row['id']}"):
                    if edited_name.strip() and edited_name.strip() != row["name"]:
                        update_employee_name(row["id"], edited_name.strip())
                    if edited_division != row["division"]:
                        update_employee_division(row["id"], edited_division)
                    if edited_order != row["order"]:
                        update_employee_order(row["id"], edited_order)
                    st.rerun()
                if cdel.button("\U0001F5D1\uFE0F Hapus", key=f"del_{row['id']}"):
                    delete_employee(row["id"])
                    st.rerun()

                replacement_photo = st.file_uploader(
                    "Ganti foto", type=["jpg", "jpeg", "png", "webp"], key=f"photofile_{row['id']}"
                )
                if replacement_photo is not None and st.button("Simpan foto ini", key=f"savephoto_{row['id']}"):
                    try:
                        with st.spinner("Memproses foto..."):
                            new_photo_url = process_uploaded_photo(replacement_photo)
                        update_employee_photo(row["id"], new_photo_url)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Gagal memproses foto: {e}")

    render_employee_management_panel(employees_df, division_options)

    st.divider()
    if st.button("\U0001F504 Refresh data dari Sheets"):
        _invalidate_employee_cache()
        _invalidate_attendance_cache()
        st.rerun()

# --- Header ---
st.markdown(
    f"""
    <div class="board-title">
        <h1>{board_title}</h1>
        <h3>{board_subtitle}</h3>
    </div>
    <div class="legend-box">
        <span class="legend-item"><span class="legend-dot" style="background:{STATUS_COLORS['merah']}"></span> &gt; 08.00 (Telat)</span>
        <span class="legend-item"><span class="legend-dot" style="background:{STATUS_COLORS['hijau']}"></span> &le; 08.00 (Tepat waktu)</span>
        <span class="legend-item"><span class="legend-dot" style="background:{STATUS_COLORS['kuning']}"></span> Tidak hadir (sakit, cuti, SPD, izin)</span>
    </div>
    """,
    unsafe_allow_html=True,
)

selected_date = st.date_input("Tanggal", value=date.today())
date_str = selected_date.isoformat()

with st.expander("\U0001F4E5 Download Rekap Bulanan"):
    exc1, exc2, exc3 = st.columns([2, 2, 1])
    ex_year = exc1.number_input("Tahun", min_value=2020, max_value=2100, value=date.today().year, step=1)
    ex_month = exc2.selectbox(
        "Bulan", options=list(range(1, 13)),
        format_func=lambda m: MONTH_NAMES_ID[m - 1],
        index=date.today().month - 1,
    )
    if exc3.button("Buat Rekap"):
        report_df = build_monthly_report(int(ex_year), int(ex_month))
        st.session_state["_last_report"] = report_df
        st.session_state["_last_report_label"] = f"{ex_year}-{int(ex_month):02d}"

    last_report = st.session_state.get("_last_report")
    if last_report is not None:
        if last_report.empty:
            st.info("Belum ada data kehadiran di bulan yang dipilih.")
        else:
            label = st.session_state.get("_last_report_label", "rekap")
            excel_bytes = build_excel_bytes(last_report)
            st.download_button(
                "\U0001F4C4 Download Excel (.xlsx)",
                data=excel_bytes,
                file_name=f"rekap_kehadiran_{label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.dataframe(last_report, use_container_width=True)

autosave_if_switching_date(date_str)


@st.fragment
def render_attendance_board(date_str, employees_df, known_divisions, division_cols_config, show_names):
    attendance_map, dirty_ids = get_working_attendance(date_str)

    if dirty_ids:
        col_warn, col_save = st.columns([4, 1])
        col_warn.warning(
            f"\u26A0\uFE0F Ada {len(dirty_ids)} perubahan kehadiran yang belum disimpan ke Google Sheets."
        )
        if col_save.button("\U0001F4BE Simpan Sekarang", type="primary", key=f"savebtn_{date_str}"):
            save_pending_attendance(date_str, {eid: attendance_map[eid] for eid in dirty_ids})
            st.session_state[f"att_dirty_{date_str}"] = set()
            st.success("Tersimpan!")
            st.rerun()

    # --- Summary ---
    counts = {"hijau": 0, "merah": 0, "kuning": 0, "belum": 0}
    for _, row in employees_df.iterrows():
        status = attendance_map.get(str(row["id"]), "belum")
        counts[status] += 1

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("\U0001F7E2 Tepat waktu", counts["hijau"])
    c2.metric("\U0001F534 Telat", counts["merah"])
    c3.metric("\U0001F7E1 Tidak hadir", counts["kuning"])
    c4.metric("\u26AA Belum diisi", counts["belum"])

    st.divider()

    if employees_df.empty:
        st.info("Belum ada karyawan. Tambahkan lewat panel di sebelah kiri.")
        return

    def render_employee_card(emp, selections, photo_size=44):
        emp_id = str(emp["id"])
        current_status = attendance_map.get(emp_id, "belum")
        photo = emp["photo_url"] if emp.get("photo_url") else placeholder_photo_url(emp["id"])
        ring = STATUS_COLORS[current_status]

        st.markdown(
            f'<div class="emp-photo-wrap"><img src="{photo}" style="width:{photo_size}px;'
            f'height:{photo_size}px;object-fit:cover;border-radius:50%;'
            f'border:2.5px solid {ring};"></div>',
            unsafe_allow_html=True,
        )
        if show_names:
            st.markdown(f'<div class="emp-row-name">{emp["name"]}</div>', unsafe_allow_html=True)
        chosen = st.selectbox(
            "Status", options=STATUS_CYCLE, index=STATUS_CYCLE.index(current_status),
            format_func=lambda s: f"{STATUS_ICON[s]} {STATUS_LABEL[s]}",
            key=f"sel_{emp_id}_{date_str}", label_visibility="collapsed",
        )
        # TIDAK langsung disimpan/rerun di sini - ditampung dulu di 'selections',
        # baru diproses sekaligus saat tombol "Terapkan" di form ditekan.
        selections[emp_id] = chosen

    selections = {}
    with st.form(f"grid_form_{date_str}", clear_on_submit=False):
        top_submit = st.form_submit_button(
            "\u2705 Terapkan Perubahan Status", key=f"submit_top_{date_str}"
        )

        kacab_df = employees_df[employees_df["division"].str.lower() == KACAB_NAME.lower()].sort_values("order")
        if not kacab_df.empty:
            st.markdown(f'<div class="kacab-header">{KACAB_NAME}</div>', unsafe_allow_html=True)
            pad = 2
            total_slots = len(kacab_df) + pad * 2
            kacab_cols = st.columns(total_slots)
            for i, (_, emp) in enumerate(kacab_df.iterrows()):
                with kacab_cols[pad + i]:
                    render_employee_card(emp, selections, photo_size=64)
            st.divider()

        grid_df = employees_df[employees_df["division"].str.lower() != KACAB_NAME.lower()]
        data_divisions = list(grid_df["division"].unique())
        ordered_divisions = [d for d in known_divisions if d in data_divisions]
        ordered_divisions += [d for d in data_divisions if d not in ordered_divisions]

        width_ratios = [division_cols_config.get(d, 1) for d in ordered_divisions]
        division_blocks = st.columns(width_ratios) if ordered_divisions else []

        for div_block, division in zip(division_blocks, ordered_divisions):
            with div_block:
                st.markdown(f'<div class="division-header">{division}</div>', unsafe_allow_html=True)
                n_cols = division_cols_config.get(division, 1)
                div_employees = grid_df[grid_df["division"] == division].sort_values("order")
                emp_list = list(div_employees.iterrows())
                chunks = chunk_column_major(emp_list, n_cols)

                sub_cols = st.columns(n_cols)
                for sub_col, chunk in zip(sub_cols, chunks):
                    with sub_col:
                        for _, emp in chunk:
                            render_employee_card(emp, selections, photo_size=40)

        bottom_submit = st.form_submit_button(
            "\u2705 Terapkan Perubahan Status", key=f"submit_bottom_{date_str}"
        )

    if top_submit or bottom_submit:
        changed = 0
        for emp_id, chosen in selections.items():
            if attendance_map.get(emp_id, "belum") != chosen:
                attendance_map[emp_id] = chosen
                dirty_ids.add(emp_id)
                changed += 1
        if changed:
            st.toast(f"{changed} status diperbarui - klik 'Simpan Sekarang' di atas untuk kirim ke Sheets.")
            st.rerun()
        else:
            st.toast("Tidak ada perubahan.")


render_attendance_board(date_str, employees_df, known_divisions, division_cols_config, show_names)

st.caption(
    "Pilih status lewat dropdown, ganti sebanyak yang kamu mau (1 divisi atau semuanya), "
    "baru klik 'Terapkan Perubahan Status' - reload cuma terjadi sekali di situ, bukan tiap "
    "ganti 1 dropdown. Setelah diterapkan, klik 'Simpan Sekarang' untuk kirim ke Google Sheets "
    "(atau otomatis tersimpan saat ganti tanggal). Kalau tutup tab sebelum simpan, perubahan bisa hilang."
)
